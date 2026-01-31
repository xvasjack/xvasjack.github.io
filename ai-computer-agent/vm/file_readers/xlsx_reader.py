"""
Excel Reader - Analyze XLSX files for quality checking.

This module extracts:
- Sheet information
- Company data from tables
- Data validation
- Missing/invalid fields

Security:
- SEC-2: ZIP bomb protection via size ratio check
"""

import os
import re
import zipfile
from typing import List, Optional, Dict, Any
from dataclasses import dataclass, field
import logging

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logging.getLogger("xlsx_reader").warning("Missing openpyxl. Run: pip install openpyxl")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("xlsx_reader")

# SEC-2: ZIP bomb protection constants
MAX_XLSX_SIZE = 50 * 1024 * 1024  # 50MB compressed
MAX_DECOMPRESSED_SIZE = 500 * 1024 * 1024  # 500MB max decompressed
MAX_COMPRESSION_RATIO = 100  # Max 100:1 compression ratio
MAX_ZIP_ENTRIES = 10000  # Max number of files in archive

# IV-6: Maximum field length to prevent memory exhaustion
MAX_FIELD_LENGTH = 10000


def _check_zip_bomb(file_path: str) -> tuple:
    """
    SEC-2: Check for ZIP bomb before processing.
    Returns (is_safe, error_message).
    """
    try:
        compressed_size = os.path.getsize(file_path)
        if compressed_size == 0:
            return False, "Empty file"

        # Check file size limit
        if compressed_size > MAX_XLSX_SIZE:
            return False, f"File too large: {compressed_size / 1024 / 1024:.1f}MB > {MAX_XLSX_SIZE / 1024 / 1024:.0f}MB"

        with zipfile.ZipFile(file_path, 'r') as zf:
            # Check number of entries
            if len(zf.namelist()) > MAX_ZIP_ENTRIES:
                return False, f"Too many files in archive: {len(zf.namelist())} > {MAX_ZIP_ENTRIES}"

            # Calculate total decompressed size
            total_uncompressed = sum(info.file_size for info in zf.infolist())

            # Check absolute size limit
            if total_uncompressed > MAX_DECOMPRESSED_SIZE:
                return False, f"Decompressed size too large: {total_uncompressed / 1024 / 1024:.1f}MB > {MAX_DECOMPRESSED_SIZE / 1024 / 1024:.0f}MB"

            # Check compression ratio
            ratio = total_uncompressed / compressed_size if compressed_size > 0 else 0
            if ratio > MAX_COMPRESSION_RATIO:
                return False, f"Suspicious compression ratio: {ratio:.1f}:1 > {MAX_COMPRESSION_RATIO}:1"

            return True, None
    except zipfile.BadZipFile:
        return False, "Invalid or corrupted ZIP file"
    except Exception as e:
        return False, f"ZIP validation error: {e}"


@dataclass
class SheetInfo:
    name: str
    row_count: int
    column_count: int
    headers: List[str]
    sample_rows: List[Dict[str, Any]]


@dataclass
class DataIssue:
    sheet: str
    row: int
    column: str
    issue: str
    value: Any


@dataclass
class XLSXAnalysis:
    file_path: str
    sheet_count: int
    sheets: List[SheetInfo] = field(default_factory=list)
    total_rows: int = 0
    issues: List[DataIssue] = field(default_factory=list)
    summary: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "file_path": self.file_path,
            "sheet_count": self.sheet_count,
            "total_rows": self.total_rows,
            "issue_count": len(self.issues),
            "sheets": [
                {
                    "name": s.name,
                    "row_count": s.row_count,
                    "column_count": s.column_count,
                    "headers": s.headers,
                }
                for s in self.sheets
            ],
            "issues": [
                {
                    "sheet": i.sheet,
                    "row": i.row,
                    "column": i.column,
                    "issue": i.issue,
                }
                for i in self.issues[:20]  # Limit issues in output
            ],
            "summary": self.summary,
        }


def analyze_xlsx(file_path: str) -> XLSXAnalysis:
    """
    Analyze an Excel file and extract relevant information.

    Args:
        file_path: Path to the XLSX file

    Returns:
        XLSXAnalysis with extracted information
    """
    logger.info(f"Analyzing XLSX: {file_path}")

    if not HAS_OPENPYXL:
        return XLSXAnalysis(
            file_path=file_path,
            sheet_count=0,
            issues=[DataIssue("", 0, "", "openpyxl not installed", None)],
            summary="Error: openpyxl not installed"
        )

    if not os.path.exists(file_path):
        return XLSXAnalysis(
            file_path=file_path,
            sheet_count=0,
            issues=[DataIssue("", 0, "", "File not found", None)],
            summary="Error: File not found"
        )

    # SEC-2: Check for ZIP bomb before processing
    is_safe, bomb_error = _check_zip_bomb(file_path)
    if not is_safe:
        logger.warning(f"ZIP bomb protection triggered: {bomb_error}")
        return XLSXAnalysis(
            file_path=file_path,
            sheet_count=0,
            issues=[DataIssue("", 0, "", f"Security: {bomb_error}", None)],
            summary=f"Security error: {bomb_error}"
        )

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return XLSXAnalysis(
            file_path=file_path,
            sheet_count=0,
            issues=[DataIssue("", 0, "", f"Could not open file: {e}", None)],
            summary=f"Error opening file: {e}"
        )

    try:
        analysis = XLSXAnalysis(
            file_path=file_path,
            sheet_count=len(wb.sheetnames),
        )

        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info = extract_sheet_info(ws, sheet_name)
            analysis.sheets.append(sheet_info)
            total_rows += sheet_info.row_count

            # Check for issues in this sheet
            sheet_issues = check_sheet_issues(ws, sheet_name, sheet_info.headers)
            analysis.issues.extend(sheet_issues)

        analysis.total_rows = total_rows
        analysis.summary = generate_summary(analysis)

        return analysis
    finally:
        wb.close()


def extract_sheet_info(ws, sheet_name: str) -> SheetInfo:
    """Extract information from a single sheet"""

    # Get dimensions
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # Category 6 fix: Empty sheets should be flagged, not pass silently
    if max_row == 0 or max_col == 0:
        logger.warning(f"Sheet '{sheet_name}' appears to be empty (rows={max_row}, cols={max_col})")
        return SheetInfo(
            name=sheet_name,
            row_count=0,
            column_count=0,
            headers=[],
            sample_rows=[],
        )

    # Get headers from first row
    headers = []
    if max_row > 0:
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=1, column=col).value
            # IV-6: Truncate field value to prevent memory exhaustion
            header_str = str(cell_value)[:MAX_FIELD_LENGTH] if cell_value else f"Column_{col}"
            headers.append(header_str)

    # Get sample rows (first 5 data rows)
    sample_rows = []
    for row in range(2, min(7, max_row + 1)):
        row_data = {}
        # IV-9: Use min() to prevent off-by-one column access
        for col in range(1, min(max_col + 1, len(headers) + 1)):
            header = headers[col - 1] if col <= len(headers) else f"Column_{col}"
            cell_val = ws.cell(row=row, column=col).value
            # IV-6: Truncate field value to prevent memory exhaustion
            if cell_val is not None:
                row_data[header] = str(cell_val)[:MAX_FIELD_LENGTH] if len(str(cell_val)) > MAX_FIELD_LENGTH else cell_val
            else:
                row_data[header] = cell_val
        sample_rows.append(row_data)

    return SheetInfo(
        name=sheet_name,
        row_count=max_row - 1 if max_row > 0 else 0,  # Exclude header
        column_count=max_col,
        headers=headers,
        sample_rows=sample_rows,
    )


def check_sheet_issues(ws, sheet_name: str, headers: List[str]) -> List[DataIssue]:
    """Check for data quality issues in a sheet"""

    issues = []

    # Common expected columns for company data
    expected_columns = {
        "company": ["company", "company name", "name", "company_name"],
        "website": ["website", "url", "web", "site", "company website"],
        "country": ["country", "location", "region"],
        "industry": ["industry", "sector", "business"],
    }

    # Normalize headers for matching
    headers_lower = [h.lower().strip() for h in headers]

    # Check if expected columns exist
    for col_type, patterns in expected_columns.items():
        found = any(any(p in h for p in patterns) for h in headers_lower)
        if not found:
            issues.append(DataIssue(
                sheet=sheet_name,
                row=1,
                column="",
                issue=f"Missing expected column: {col_type}",
                value=None
            ))

    # Check each row for issues
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    for row in range(2, max_row + 1):
        row_values = [ws.cell(row=row, column=col).value for col in range(1, max_col + 1)]

        # Check for empty rows
        if all(v is None or str(v).strip() == "" for v in row_values):
            issues.append(DataIssue(
                sheet=sheet_name,
                row=row,
                column="all",
                issue="Empty row",
                value=None
            ))
            continue

        # Check each cell
        # Category 11 fix: Extra columns silently dropped by zip() - iterate all columns
        for col_idx in range(1, max(len(headers), len(row_values)) + 1):
            header = headers[col_idx - 1] if col_idx <= len(headers) else f"Column_{col_idx}"
            value = row_values[col_idx - 1] if col_idx <= len(row_values) else None
            header_lower = header.lower()

            # Check for empty required fields
            if any(p in header_lower for p in ["company", "name"]):
                if value is None or str(value).strip() == "":
                    issues.append(DataIssue(
                        sheet=sheet_name,
                        row=row,
                        column=header,
                        issue="Empty company name",
                        value=value
                    ))

            # Check website format
            if any(p in header_lower for p in ["website", "url", "web"]):
                if value and not is_valid_url(str(value)):
                    issues.append(DataIssue(
                        sheet=sheet_name,
                        row=row,
                        column=header,
                        issue="Invalid URL format",
                        value=value
                    ))

            # Check for placeholder values
            if value and is_placeholder_value(str(value)):
                issues.append(DataIssue(
                    sheet=sheet_name,
                    row=row,
                    column=header,
                    issue="Placeholder value detected",
                    value=value
                ))

    return issues


def is_valid_url(url: str) -> bool:
    """Basic URL validation"""
    if not url:
        return False

    url = url.strip().lower()

    # Check for common URL patterns
    if url.startswith(("http://", "https://", "www.")):
        return True

    # Check for domain-like pattern
    if re.match(r'^[a-z0-9][-a-z0-9]*\.[a-z]{2,}', url):
        return True

    return False


def is_placeholder_value(value: str) -> bool:
    """Check if value looks like a placeholder"""
    placeholders = [
        "n/a", "na", "none", "null", "undefined",
        "tbd", "tbc", "xxx", "---", "...",
        "placeholder", "test", "example",
    ]

    value_lower = value.lower().strip()

    return value_lower in placeholders


def generate_summary(analysis: XLSXAnalysis) -> str:
    """Generate a summary of the analysis"""

    sheet_info = "\n".join(
        f"  - {s.name}: {s.row_count} rows, {s.column_count} columns"
        for s in analysis.sheets
    )

    issue_summary = {}
    for issue in analysis.issues:
        issue_type = issue.issue
        issue_summary[issue_type] = issue_summary.get(issue_type, 0) + 1

    issue_breakdown = "\n".join(
        f"  - {issue_type}: {count} occurrences"
        for issue_type, count in sorted(issue_summary.items(), key=lambda x: -x[1])
    )

    summary = f"""XLSX Analysis Summary:
- File: {os.path.basename(analysis.file_path)}
- Sheets: {analysis.sheet_count}
- Total data rows: {analysis.total_rows}
- Issues found: {len(analysis.issues)}

Sheets:
{sheet_info}

Issue breakdown:
{issue_breakdown if issue_breakdown else "  - No issues found"}
"""

    return summary


# =============================================================================
# DATA EXTRACTION
# =============================================================================


def extract_companies_from_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """
    Extract company data from an Excel file.

    Returns list of company dictionaries.
    """
    logger.info(f"Extracting companies from: {file_path}")

    if not HAS_OPENPYXL:
        logger.error("openpyxl not installed")
        return []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        logger.error(f"Could not open file: {e}")
        return []

    try:
        companies = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get headers
            headers = []
            max_col = ws.max_column or 0
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=1, column=col).value
                headers.append(str(cell_value).lower().strip() if cell_value else f"col_{col}")

            # Map headers to standard fields
            field_mapping = {}
            for idx, header in enumerate(headers):
                if any(p in header for p in ["company", "name"]):
                    field_mapping["name"] = idx
                elif any(p in header for p in ["website", "url", "web"]):
                    field_mapping["website"] = idx
                elif any(p in header for p in ["country", "location"]):
                    field_mapping["country"] = idx
                elif any(p in header for p in ["industry", "sector"]):
                    field_mapping["industry"] = idx
                elif any(p in header for p in ["description", "desc", "about"]):
                    field_mapping["description"] = idx

            # Extract rows
            max_row = ws.max_row or 0
            for row in range(2, max_row + 1):
                company = {}

                for field, col_idx in field_mapping.items():
                    value = ws.cell(row=row, column=col_idx + 1).value
                    company[field] = value

                # Only add if has a name
                if company.get("name"):
                    company["source_sheet"] = sheet_name
                    company["source_row"] = row
                    companies.append(company)

        return companies
    finally:
        wb.close()


# =============================================================================
# COMPARISON FUNCTIONS
# =============================================================================


def compare_analyses(old: XLSXAnalysis, new: XLSXAnalysis) -> Dict[str, Any]:
    """
    Compare two XLSX analyses to see if issues were fixed.

    Returns dict with comparison results.
    """
    old_issue_types = {i.issue for i in old.issues}
    new_issue_types = {i.issue for i in new.issues}

    return {
        "row_count_change": new.total_rows - old.total_rows,
        "old_issue_count": len(old.issues),
        "new_issue_count": len(new.issues),
        "issues_fixed_types": list(old_issue_types - new_issue_types),
        "new_issue_types": list(new_issue_types - old_issue_types),
        "improved": len(new.issues) < len(old.issues),
    }


# =============================================================================
# ENTRY POINT
# =============================================================================


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python xlsx_reader.py <file.xlsx>")
        sys.exit(1)

    analysis = analyze_xlsx(sys.argv[1])
    print(analysis.summary)
